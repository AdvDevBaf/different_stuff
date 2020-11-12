import pandas as pd
import requests
import json
from bs4 import BeautifulSoup
import urllib
from tkinter import *
from tkinter.filedialog import askdirectory,askopenfilename
import openpyxl

from os import listdir, path
class Gui(Toplevel):
    def __init__(self, parent, title="Обработка файлов"):
        Toplevel.__init__(self, parent)
        parent.geometry("250x250+100+150")
        if title:
            self.title(title)
        parent.withdraw()
        self.parent = parent
        self.result = None
        dialog = Frame(self)
        self.initial_focus = self.dialog(dialog)
        self.protocol("WM_DELETE_WINDOW", self.on_exit)
        dialog.pack()

    def on_exit(self):
        self.quit()

    def text_3_on(self):
        if self.var_1.get():
            self.text_3["state"] = "normal"
            self.text_3.delete(0, END)
            self.text_3.insert(END, "Name_{{initial_image_name}}")
        else:
            self.text_3["state"] = "disabled"

    def search_folder_for_files(self):
        path_to = askopenfilename()
        print(path_to)
        self.text_1.delete(0, END)
        self.text_1.insert(END, path_to)

    def search_folder_for_new_excel_file(self):
        path_to = askdirectory()
        print(path_to)
        self.text_2.delete(0, END)
        self.text_2.insert(END, path_to)

    def save_to_csv(self,tittle, date_of_publish, comments):

        table = pd.DataFrame({'Имя канала': tittle, 'Дата публикации комментария': date_of_publish,
                              'Текст': comments})
        if self.var_1.get():
            table.to_csv(str(self.text_2.get()) + '/' + str(self.text_3.get()) + '.csv', sep=';', index=False,encoding='utf-8-sig')
        else:
            table.to_csv(str(self.text_2.get()) + '/' + 'Комментарии' + '.csv', sep=';', index=False,encoding='utf-8-sig')
        import ctypes
        message = 'Готово!'
        ctypes.windll.user32.MessageBoxW(0, message, 'Получить комментарии', 0)
        print('ok')
        return {}

    def get_comments(self, uploads):
        videoId = []
        comments = []
        date_of_publish = []
        errors = []
        ch_id = []
        tittle = []

        # Получим список видеозаписей, к которым нужно получить комментарии
        for name in uploads:

            response = requests.get('https://www.googleapis.com/youtube/v3/playlistItems?playlistId=' + str(name)
                                    + '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&maxResults=50&'
                                      'part=snippet,contentDetails')
            last_upload = response.json()
            for i in range(0, len(last_upload['items'])):
                videoId.append(last_upload["items"][i]["snippet"]["resourceId"]["videoId"])
                print(last_upload["items"][i]["snippet"]["resourceId"]["videoId"])
                print('ne token')
            #try:
            if 'nextPageToken' in last_upload:
                print('token?')
                print(last_upload)
                for i in range(0, (int(last_upload['pageInfo']["totalResults"])//50)+1):
                    if 'nextPageToken' in last_upload:
                        print('token??')
                        response = requests.get('https://www.googleapis.com/youtube/v3/playlistItems?playlistId=' +
                                                str(name) + '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&pageToken=' +
                                                str(last_upload['nextPageToken']) + '&maxResults=50'
                                                                                    '&part=snippet,contentDetails')
                        last_upload = response.json()
                        print('token')
                        for j in range(0, len(last_upload['items'])):
                            videoId.append(last_upload["items"][j]["snippet"]["resourceId"]["videoId"])
                            print(last_upload["items"][j]["snippet"]["resourceId"]["videoId"])
                            print('oh token')
            #except:
             #   errors.append(IndexError.args)

        print(len(videoId))

    # Получим список комментариев и дату их создания

        for vid in videoId:
            response = requests.get('https://www.googleapis.com/youtube/v3/commentThreads?videoId=' + str(vid)
                                    + '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&part=snippet&maxResults=100')
            comms = response.json()
            print(comms)
            # Получим имена каналов
            response = requests.get('https://www.googleapis.com/youtube/v3/videos?id=' + str(vid)
                                    + '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&part=snippet')
            data = response.json()

            try:
                for i in range(1, len(comms['items'])):
                    print(comms["items"][i]["snippet"]["topLevelComment"]["snippet"]["textDisplay"])
                    comments.append(comms["items"][i]["snippet"]["topLevelComment"]["snippet"]["textDisplay"])
                    date_of_publish.append(comms["items"][i]["snippet"]["topLevelComment"]["snippet"]["publishedAt"])
                    ch_id.append(data["items"][0]["snippet"]['channelId'])  # Имя канала
                    print('ne token')

                if 'nextPageToken' in comms:
                    print(comms)
                    try:
                        for i in range(0, int(comms['pageInfo']["totalResults"]) // 100 + 1):
                            if 'nextPageToken' in comms:
                                print('token?')
                                response = requests.get('https://www.googleapis.com/youtube/v3/commentThreads?videoId=' +
                                                    str(vid) + '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU'
                                                               '&part=snippet&maxResults=100&pageToken=' +
                                                    str(comms['nextPageToken']))
                                comms = response.json()
                                print('token da')
                                for j in range(0, len(comms['items'])):
                                    print(comms["items"][j]["snippet"]["topLevelComment"]["snippet"]["textDisplay"])
                                    comments.append(comms["items"][j]["snippet"]["topLevelComment"]["snippet"]["textDisplay"])
                                    date_of_publish.append(comms["items"][j]["snippet"]["topLevelComment"]["snippet"]["publishedAt"])
                                    ch_id.append(data["items"][0]["snippet"]['channelId'])  # Имя канала
                                    print('tok-token')
                    except:
                        print(i)
                        print('Нет следующей страницы')
            except:
                print(i)
                print("Нет комментариев")

        # Получим названия каналов, для которых требуется получить комментарии
        for name in ch_id:
            content = requests.get('https://www.googleapis.com/youtube/v3/channels?id=' + str(
                name) + '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&part=snippet,contentDetails')
            data = content.json()
            tittle.append(data["items"][0]["snippet"]["title"])

        print(len(ch_id))
        print(len(data))
        print(len(comments))
        print(len(date_of_publish))

        return comments, date_of_publish, tittle

    def get_channel_name_list(self):
        names = []
        wb = openpyxl.load_workbook(str(self.text_1.get()).replace('/', '\\'))
        sheet = wb.worksheets[0]
        for i in range(1, sheet.max_row):
            if ((sheet.cell(row=i, column=1).value) == None):
                max_row = i - 1
                break
            else:
                max_row = sheet.max_row

        for i in range(2, max_row + 1):
            names.append(str(sheet.cell(row=i, column=1).value)[28:])
        print(names)
        print('ok')
        return names

    def start(self):
        channel_name = self.get_channel_name_list()
        uploads = []

        for name in channel_name:
            content = requests.get('https://www.googleapis.com/youtube/v3/channels?id=' + str(
                name) + '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&part=snippet,contentDetails')
            data = content.json()
            try:
                print(data["items"][0]["contentDetails"]["relatedPlaylists"]["uploads"])
                uploads.append(data["items"][0]["contentDetails"]["relatedPlaylists"]["uploads"])
            except:
                print('Канал удален')

        comments, date_of_publish, tittle = self.get_comments(uploads)

        return self.save_to_csv(tittle, date_of_publish, comments)


    def dialog(self, parent):
        self.parent = parent

        # Created main elements
        self.label_1 = Label(parent, text="Укажите путь, по которому лежит основной Excel файл")
        self.text_1 = Entry(parent, width=50)
        self.but_1 = Button(parent, text="Указать", command=self.search_folder_for_files)

        self.var_1 = IntVar()

        self.label_2 = Label(parent, text="Укажие папку, куда положить получившуюся таблицу")
        self.text_2 = Entry(parent, width=50)
        self.but_2 = Button(parent, text="Указать", command=self.search_folder_for_new_excel_file)

        self.chk_1 = Checkbutton(parent, text="Переименовать файл по маске", variable=self.var_1, command=self.text_3_on)
        self.text_3 = Entry(parent, width=50, state=DISABLED, disabledforeground=parent.cget('bg'))

        self.label_1.pack()
        self.text_1.pack()
        self.but_1.pack()

        self.label_2.pack()
        self.text_2.pack()
        self.but_2.pack()

        self.chk_1.pack()
        self.text_3.pack()

        # start button
        self.but_start = Button(parent, text="Выполнить",command=self.start)
        self.but_start.pack()


if __name__ == "__main__":
        root = Tk()
        root.minsize(width=500, height=400)
        gui = Gui(root)
        root.mainloop()

